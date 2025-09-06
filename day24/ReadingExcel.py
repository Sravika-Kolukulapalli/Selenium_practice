import openpyxl

#File----->Workbook--->Sheet--->Rows--->Cells
file="/Users/srakoluk/Downloads/Employee_Details.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
rows=sheet.max_row  #gives number of rows in excel sheet
cols=sheet.max_column #gives number of columns in excel sheet


#Read data from Excel
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="     ")
    print()




