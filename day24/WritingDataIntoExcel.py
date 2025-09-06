# import openpyxl
#
#
# file="/Users/srakoluk/Downloads/Empty_Excel_Sheet.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook.active      #(or)    #workbook["Sheet"]----get active sheet from excel
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Hiiiiiiii"



############################################
#multiple data
import openpyxl


file="/Users/srakoluk/Downloads/testdata1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active



sheet.cell(1,1).value="1"
sheet.cell(1,2).value="Sravs"
sheet.cell(1,3).value="Software Engineer"


sheet.cell(2,1).value="2"
sheet.cell(2,2).value="Raj"
sheet.cell(2,3).value="Security Gaurd"


sheet.cell(3,1).value="3"
sheet.cell(3,2).value="Deepak"
sheet.cell(3,3).value="Sweeper"

workbook.save(file)#save file after entering data

